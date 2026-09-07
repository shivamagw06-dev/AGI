#!/usr/bin/env python3
"""Build a Capital IQ Excel plug-in workbook for estimate vintages.

Produces a workbook where every data cell is a live CIQ() formula wired to a
ticker column and a date row. Paste identifiers into column B and the plug-in
resolves the grid on open.

Design decisions worth knowing:

* Every mnemonic lives in ONE cell on the Setup sheet and all formulas
  reference it. Capital IQ mnemonics vary by entitlement and version, so a
  wrong guess must be a one-cell correction, not a rebuild of 30,000 formulas.
* Date headers are computed with EOMONTH from a single start date, so shifting
  the whole window is one edit.
* Formulas are wrapped in IFERROR so an unentitled field renders blank instead
  of papering the sheet with #NAME?.
* A period-end-date column sits beside the estimate grid. FY1 is relative to
  the as-of date, so without it you cannot tell which fiscal year a cell means
  and the series is unusable for anything but a relative factor.

    python3 scripts/build_capiq_template.py [-o OUT.xlsx] [--rows 520] [--months 72]
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INK = "1F5F5B"
HEAD = "E8EDEC"
WARN = "FDF3E3"
THIN = Side(style="thin", color="D6D3CC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(ws, cell: str, text: str, size: int = 13):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=INK)


def _head(ws, row: int, col: int, text: str):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
    return c


def build_setup(wb) -> None:
    ws = wb.create_sheet("Setup", 0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 78

    _title(ws, "A1", "Capital IQ vintage pull — setup", 15)
    ws["A2"] = "Change values in column B only. Every formula on the other sheets points here."
    ws["A2"].font = Font(italic=True, size=10)

    rows = [
        ("", "", ""),
        ("MNEMONICS", "", "Verify each in the CIQ ribbon → Formula Builder before pulling at scale."),
        ("EPS estimate (mean)", "IQ_EPS_EST",
         "Search 'EPS Estimate' in Formula Builder and paste the exact mnemonic here."),
        ("Revenue estimate", "IQ_REVENUE_EST", "Optional. Less manipulable than EPS."),
        ("Number of estimates", "IQ_NUM_EST", "Coverage. Thin coverage is noise, not signal."),
        ("Estimate std deviation", "IQ_EST_STDDEV", "Dispersion is a factor in its own right."),
        ("Period end date", "IQ_PERIODDATE",
         "CRITICAL — tells you which fiscal year each cell refers to. See note below."),
        ("Actual reported EPS", "IQ_DILUT_EPS_EXCL",
         "Optional but high value: estimate + actual = earnings surprise."),
        ("", "", ""),
        ("PARAMETERS", "", ""),
        ("Period", "FY1", "FY1 = one year forward. FY2 for two."),
        ("Window start (month-end)", "=DATE(2020,1,31)",
         "First column date. All other columns step forward one month from here."),
        ("Currency", "INR", "Leave blank to use the company's reporting currency."),
        ("", "", ""),
        ("HOW TO USE", "", ""),
        ("1.", "", "Enable the Capital IQ plug-in and sign in."),
        ("2.", "", "Paste ISINs into column A and CIQ identifiers into column B of 'Vintages_EPS'."),
        ("3.", "", "Identifier format: TICKER:EXCHANGE, e.g. RELIANCE:NSEI. ISIN also works."),
        ("4.", "", "TEST FIRST: keep 20 rows and 12 columns, confirm numbers return, then scale."),
        ("5.", "", "When populated: Copy → Paste Special → Values, then save. "
                   "Live formulas will not resolve outside your machine."),
        ("", "", ""),
        ("WHY THE PERIOD-END COLUMN MATTERS", "", ""),
        ("", "", "FY1 is relative to the as-of date, not fixed. On 31-Mar-2023 FY1 means FY2023; "
                 "on 31-Mar-2025 it means FY2025. A row of FY1 values is therefore a moving target — "
                 "correct for a one-year-forward factor, but not comparable as levels. The period-end "
                 "column records which fiscal year each cell actually refers to."),
        ("", "", ""),
        ("IF CELLS COME BACK BLANK", "", ""),
        ("", "", "Recent dates working and older dates blank means estimate history is not in your "
                 "entitlement — it is often licensed separately from current estimates. Find that out "
                 "at 20 rows, not after building 30,000 cells."),
    ]
    r = 3
    for label, value, note in rows:
        ws.cell(row=r, column=1, value=label).font = Font(
            bold=label.isupper() and bool(label), size=10)
        if value:
            c = ws.cell(row=r, column=2, value=value)
            c.font = Font(bold=True, size=10, name="Consolas")
            c.fill = PatternFill("solid", fgColor=WARN)
            c.border = BOX
        n = ws.cell(row=r, column=3, value=note)
        n.font = Font(size=9, color="5A5F63")
        n.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    ws.freeze_panes = "A3"


# Setup cell references used by every formula.
M_EPS, M_REV, M_NUM, M_STD, M_PERIOD, M_ACTUAL = (
    "Setup!$B$5", "Setup!$B$6", "Setup!$B$7", "Setup!$B$8", "Setup!$B$9", "Setup!$B$10")
P_PERIOD, P_START, P_CCY = "Setup!$B$13", "Setup!$B$14", "Setup!$B$15"


def build_grid(wb, name: str, mnemonic_ref: str, description: str,
               rows: int, months: int, first_data_col: int = 4) -> None:
    """One vintage grid: identifiers down, month-end as-of dates across."""
    ws = wb.create_sheet(name)
    _title(ws, "A1", description, 12)
    ws["A2"] = ("Paste ISIN in column A and the CIQ identifier in column B. "
                "Row 3 holds the as-of dates; each cell asks what consensus was on that date.")
    ws["A2"].font = Font(italic=True, size=9, color="5A5F63")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14

    _head(ws, 3, 1, "ISIN")
    _head(ws, 3, 2, "CIQ identifier")
    _head(ws, 3, 3, "Period end")

    for i in range(months):
        col = first_data_col + i
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 11
        # Month-ends stepping forward from the single start date on Setup.
        c = ws.cell(row=3, column=col, value=f"=EOMONTH({P_START},{i})")
        c.number_format = "dd-mmm-yy"
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center")
        c.border = BOX

    for r in range(4, 4 + rows):
        # Period end for the most recent as-of date, so the fiscal year each
        # FY1 refers to is recorded rather than inferred.
        last = get_column_letter(first_data_col + months - 1)
        pc = ws.cell(row=r, column=3,
                     value=f'=IFERROR(CIQ($B{r},{M_PERIOD},{P_PERIOD},{last}$3),"")')
        pc.number_format = "dd-mmm-yy"
        pc.border = BOX
        for i in range(months):
            col = first_data_col + i
            letter = get_column_letter(col)
            cell = ws.cell(
                row=r, column=col,
                value=f'=IFERROR(CIQ($B{r},{mnemonic_ref},{P_PERIOD},{letter}$3),"")',
            )
            cell.number_format = "0.00"
            cell.border = BOX

    ws.freeze_panes = ws.cell(row=4, column=first_data_col)


def build_cross_section(wb, rows: int) -> None:
    """Current forward estimates — the Part 1 pull, as a plug-in alternative."""
    ws = wb.create_sheet("Forward_Estimates_Now")
    _title(ws, "A1", "Forward estimates — current snapshot", 12)
    ws["A2"] = ("Alternative to the Screener export. Same fields, pulled through the plug-in. "
                "No as-of date, so this is today's consensus only.")
    ws["A2"].font = Font(italic=True, size=9, color="5A5F63")

    cols = [
        ("ISIN", None, 16), ("CIQ identifier", None, 20),
        ("EPS FY1", M_EPS, 12), ("EPS FY2", M_EPS, 12),
        ("Revenue FY1", M_REV, 14), ("Revenue FY2", M_REV, 14),
        ("# est FY1", M_NUM, 11), ("Std dev FY1", M_STD, 12),
        ("Period end FY1", M_PERIOD, 14), ("Actual EPS", M_ACTUAL, 12),
    ]
    for i, (label, _, width) in enumerate(cols, start=1):
        _head(ws, 3, i, label)
        ws.column_dimensions[get_column_letter(i)].width = width

    for r in range(4, 4 + rows):
        for i, (label, mnem, _) in enumerate(cols, start=1):
            if mnem is None:
                continue
            period = '"FY2"' if "FY2" in label else P_PERIOD
            cell = ws.cell(row=r, column=i,
                           value=f'=IFERROR(CIQ($B{r},{mnem},{period}),"")')
            cell.border = BOX
            cell.number_format = "dd-mmm-yy" if "Period" in label else "0.00"
    ws.freeze_panes = "C4"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("-o", "--out", default="capiq_vintage_template.xlsx")
    ap.add_argument("--rows", type=int, default=520, help="identifier rows (default 520)")
    ap.add_argument("--months", type=int, default=72, help="as-of columns (default 72 = 6 years)")
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)
    build_setup(wb)
    build_grid(wb, "Vintages_EPS", M_EPS,
               "EPS estimate vintages — what consensus was on each date", args.rows, args.months)
    build_grid(wb, "Vintages_NumEst", M_NUM,
               "Number of estimates — coverage at each date", args.rows, args.months)
    build_cross_section(wb, args.rows)

    out = Path(args.out)
    wb.save(out)
    cells = args.rows * args.months * 2 + args.rows * 8
    print(f"wrote {out}")
    print(f"  rows={args.rows}  months={args.months}")
    print(f"  ~{cells:,} live CIQ formulas")
    print("  sheets: Setup, Vintages_EPS, Vintages_NumEst, Forward_Estimates_Now")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
