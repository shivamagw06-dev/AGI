"""Controlled import of the repository's annual Capital IQ workbook.

The workbook is a source snapshot, not a live vendor call. It imports the full
2016–2026 supplied history and records the exact workbook and unit on every
write. This protects downstream ratio calculations from legacy Yahoo /
unknown-unit statement rows.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from financial_warehouse_completion.models import ENGINE_CODE, PROGRAMME_VERSION


SOURCE = "capital_iq_workbook"
DEFAULT_YEARS = tuple(range(2016, 2027))
_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = _ROOT / "Master_10Y_India.xlsx"
LEGACY_WORKBOOK_PATH = _ROOT / "2016-2026.xlsx"

FIELD_MAP = {
    "Revenue": "revenue", "Gross Profit": "gross_profit", "EBITDA": "ebitda",
    "EBIT": "ebit", "PBT": "pbt", "PAT": "pat", "EPS": "eps",
    "Cash & Equivalents": "cash", "Total Current Assets": "current_assets",
    "Total Assets": "assets", "Total Current Liabilities": "current_liabilities",
    "Total Debt": "debt", "Total Equity": "equity",
    "Working Capital": "working_capital", "Cash Flow from Operations": "cfo",
    "Capital Expenditure": "capex", "Free Cash Flow": "free_cash_flow",
    "Cash Flow from Investing": "cfi", "Cash Flow from Financing": "cff",
}

MASTER_FIELD_MAP = {
    "Revenue": "revenue", "EBITDA": "ebitda", "EBIT": "ebit", "Depreciation": "depreciation",
    "Amortization": "amortization", "EBITA": "ebita", "PBT": "pbt", "Tax": "tax_expense",
    "PAT": "pat", "Minority interest": "minority_interest", "Exceptional items": "exceptional_items",
    "Finance cost": "finance_cost", "R&D": "research_and_development", "Employee cost": "employee_cost",
    "Other operating exp": "other_operating_expense", "Cash": "cash",
    "ST investments": "short_term_investments", "Total investments": "total_investments",
    "Accounts receivable": "accounts_receivable", "Inventory": "inventory",
    "Other current assets": "other_current_assets", "PPE (net)": "net_ppe",
    "Intangible assets": "intangible_assets", "Goodwill": "goodwill", "Total assets": "assets",
    "Accounts payable": "accounts_payable", "Other current liab": "other_current_liabilities",
    "Total debt": "debt", "Current debt": "current_debt", "Long-term debt": "long_term_debt",
    "Lease liabilities (LT)": "lease_liabilities", "Total liabilities": "total_liabilities",
    "Equity (total common)": "equity", "CFO": "cfo", "CFI": "cfi", "CFF": "cff",
    "Capex": "capex", "Depreciation (CF)": "depreciation", "Acquisition spending": "acquisition_spending",
    "Dividends": "dividends_paid", "Buybacks": "buybacks", "Debt issuance": "debt_issuance",
    "Debt repayment": "debt_repayment",
}


def _symbol(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text.split(":", 1)[-1] if ":" in text else text


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sheet_rows(year: int, *, path: Path) -> Iterable[dict[str, Any]]:
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        if str(year) not in book.sheetnames:
            return
        sheet = book[str(year)]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
        positions = {str(label).strip(): index for index, label in enumerate(headers) if label}
        for values in sheet.iter_rows(min_row=4, values_only=True):
            symbol = _symbol(values[positions.get("Ticker", 0)] if values else None)
            if not symbol:
                continue
            row: dict[str, Any] = {
                "symbol": symbol,
                "fiscal_year": f"FY{year}",
                "statement_type": "CONSOLIDATED",
                "statement_frequency": "ANNUAL",
                "source": SOURCE,
                "statement_version": f"capiq_workbook_{year}",
            }
            for label, field in FIELD_MAP.items():
                index = positions.get(label)
                if index is not None:
                    number = _number(values[index])
                    if number is not None:
                        row[field] = number
            # A row with no financial facts is a vendor no-coverage marker.
            if any(row.get(field) is not None for field in FIELD_MAP.values()):
                yield row
    finally:
        book.close()


def _master_rows(*, path: Path) -> Iterable[dict[str, Any]]:
    """Normalize the three-statement, year-block Capital IQ export."""
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    joined: dict[tuple[str, str], dict[str, Any]] = {}
    try:
        for sheet_name in ("Income Statement", "Balance Sheet", "Cash Flow"):
            sheet = book[sheet_name]
            header = list(sheet.iter_rows(min_row=1, max_row=4, values_only=True))
            periods: dict[int, tuple[str, Any]] = {}
            current_year = None
            for index in range(1, sheet.max_column):
                if header[0][index]:
                    text = str(header[0][index])
                    current_year = text.split(" ", 1)[0]
                periods[index] = (current_year, header[2][index])
            for values in sheet.iter_rows(min_row=6, values_only=True):
                symbol = _symbol(values[0] if values else None)
                if not symbol:
                    continue
                for index in range(1, min(len(values), sheet.max_column)):
                    fiscal_year, fiscal_end = periods[index]
                    metric = str(header[1][index] or "").strip()
                    field = MASTER_FIELD_MAP.get(metric)
                    value = _number(values[index])
                    if not field or value is None or not fiscal_year:
                        continue
                    key = (symbol, fiscal_year)
                    row = joined.setdefault(key, {
                        "symbol": symbol, "fiscal_year": fiscal_year,
                        "fiscal_end_date": fiscal_end.isoformat() if hasattr(fiscal_end, "isoformat") else str(fiscal_end or "")[:10],
                        "statement_type": "UNKNOWN", "statement_frequency": "ANNUAL",
                        "source": SOURCE, "statement_version": f"capiq_master_10y_{fiscal_year.lower()}",
                    })
                    row[field] = value
        for row in joined.values():
            if row.get("cfo") is not None and row.get("capex") is not None:
                row["free_cash_flow"] = float(row["cfo"]) - abs(float(row["capex"]))
            yield row
    finally:
        book.close()


def preview(*, years: Iterable[int] = DEFAULT_YEARS, path: Path = WORKBOOK_PATH) -> dict[str, Any]:
    if not path.is_file():
        return {"ok": False, "error": f"workbook_not_found:{path.name}"}
    from openpyxl import load_workbook
    probe = load_workbook(path, read_only=True)
    sheetnames = probe.sheetnames
    probe.close()
    if "Income Statement" in sheetnames:
        rows = list(_master_rows(path=path))
        summary = {str(year): sum(1 for row in rows if row["fiscal_year"] == f"FY{year}") for year in years}
        return {"ok": True, "source": SOURCE, "workbook": path.name, "unit": "INR million", "years": summary,
                "engine": ENGINE_CODE, "version": PROGRAMME_VERSION, "format": "three_statement_year_blocks"}
    summary: dict[str, int] = {}
    for year in years:
        summary[str(year)] = sum(1 for _ in _sheet_rows(int(year), path=path))
    return {
        "ok": True, "source": SOURCE, "workbook": path.name,
        "unit": "INR million", "years": summary,
        "engine": ENGINE_CODE, "version": PROGRAMME_VERSION,
    }


def audit_preview(*, years: Iterable[int] = DEFAULT_YEARS, path: Path = WORKBOOK_PATH,
                  limit: int | None = None) -> dict[str, Any]:
    """Run identity, period and completeness gates without writing financials."""
    from financial_warehouse_completion.capiq_normalization import audit_and_prepare

    selected = tuple(sorted({int(year) for year in years if 2016 <= int(year) <= 2026}))
    check = preview(years=selected, path=path)
    if check.get("format") == "three_statement_year_blocks":
        rows = [row for row in _master_rows(path=path) if int(str(row["fiscal_year"]).replace("FY", "")) in selected]
        field_map = MASTER_FIELD_MAP
    else:
        rows = [row for year in selected for row in _sheet_rows(year, path=path)]
        field_map = FIELD_MAP
    if limit is not None:
        rows = rows[:max(0, int(limit))]
    prepared = audit_and_prepare(rows, field_map=field_map, source_file=path.name)
    audits = prepared["audits"]
    by_status: dict[str, int] = {}
    for row in audits:
        status = str(row.get("overall_status") or "UNKNOWN")
        by_status[status] = by_status.get(status, 0) + 1
    return {
        "ok": True, "source": SOURCE, "workbook": path.name, "years": list(selected),
        "seen": len(rows), "ready": len(prepared["accepted"]), "status_counts": by_status,
        "sample": audits[:25], "mapping_version": "CAPIQ_V1", "unit": "INR million",
    }


def import_completed_years(*, years: Iterable[int] = DEFAULT_YEARS, actor: str = "fwcp") -> dict[str, Any]:
    from institutional_warehouse.formulas import recalculate
    from financial_warehouse_completion.capiq_normalization import audit_and_prepare, persist

    selected = tuple(sorted({int(year) for year in years if 2016 <= int(year) <= 2026}))
    check = preview(years=selected)
    if not check.get("ok"):
        return check
    if check.get("format") == "three_statement_year_blocks":
        rows = [row for row in _master_rows(path=WORKBOOK_PATH) if int(str(row["fiscal_year"]).replace("FY", "")) in selected]
        field_map = MASTER_FIELD_MAP
    else:
        rows = [row for year in selected for row in _sheet_rows(year, path=WORKBOOK_PATH)]
        field_map = FIELD_MAP
    if not rows:
        return {"ok": False, "error": "no_financial_rows", **check}
    prepared = audit_and_prepare(rows, field_map=field_map, source_file=WORKBOOK_PATH.name)
    written = persist(prepared, field_map=field_map, actor=actor, source_file=WORKBOOK_PATH.name)
    rebuilt = recalculate(
        actor=actor,
        stages=("statement_derivations", "ratios", "annual_sector_ratios", "valuation", "factors", "quality"),
    )
    return {
        "ok": True, "source": SOURCE, "workbook": WORKBOOK_PATH.name,
        "years": list(selected), "rows": len(rows), "ready": len(prepared["accepted"]),
        "financials_annual": written["financials"], "identity": written["identity"],
        "audit": written["audit"], "metric_mapping": written["mapping"],
        "recalculated": rebuilt, "unit": "INR million", "engine": ENGINE_CODE,
        "version": PROGRAMME_VERSION,
    }
