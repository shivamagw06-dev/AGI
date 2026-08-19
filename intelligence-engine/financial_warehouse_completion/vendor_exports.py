"""Controlled import of checked-in vendor exports (Trendlyne, Capital IQ).

Three things this module exists to fix, all confirmed against production on
2026-08-19:

* ``historical_industry_medians`` and ``historical_market_medians`` are both
  empty, so relative-value scans derive medians on the fly from a thin,
  unwinsorised cross-section. That produced an "industry median" EV/EBITDA of
  48.85 for Communications Equipment and ranked BIRLACABLE as a 94% discount.
  Trendlyne publishes vendor-computed sector and industry aggregates; we store
  those instead of recomputing them.
* ``ownership`` holds 9,894 rows with no promoter-pledge signal. The Trendlyne
  multigroup export carries promoter / FII / MF / institutional holdings plus
  pledge percentage for ~3,116 companies.
* Nothing stores average daily traded value, so no strategy can size a position
  against liquidity. The Capital IQ broker export carries 3-month ADV for 2,947
  names.

Deliberately NOT handled here:

* Forward estimates. Trendlyne returns the literal string ``Export NA`` for
  ``Forecaster Estimates 1Y forward PE/PEG`` in every export produced so far,
  at both 500 and 3,486 row universes. Forward EPS must come from Capital IQ.
* ``consensus``. The broker export contains 910 covered names, which is exactly
  what the warehouse already holds — it is fully ingested, not partially.
* ``institutional_flow``. That tab is market-level daily FII/DII cash flow
  keyed on (date, segment, interval). Per-company holdings belong in
  ``ownership`` and are not a substitute for it.
"""

from __future__ import annotations

import hashlib
import statistics
import threading
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

_ROOT = Path(__file__).resolve().parents[2]

SOURCE_TRENDLYNE = "trendlyne_data_downloader"
SOURCE_CAPIQ_BROKER = "capital_iq_broker_estimates"
CODE_VERSION = "vendor_exports_v1"

_SEED_LOCK = threading.Lock()

# Vendor sentinels. Trendlyne emits "Export NA" for fields outside the plan's
# entitlement; Capital IQ emits 0 for "no data" on ratings and target price,
# which must not be confused with a genuine zero.
_TRENDLYNE_NULLS = {"", "-", "--", "na", "n/a", "nm", "n.m.", "export na", "none"}

# Trendlyne column -> (warehouse metric key, cap used to reject contaminated values)
_INDUSTRY_METRICS = {
    "Industry PE TTM": ("pe", 300.0),
    "Industry Price to Book TTM": ("pb", 60.0),
    "Industry PEG TTM": ("peg", 25.0),
    "Industry Return on Equity ROE": ("roe", 200.0),
    "Industry Return on Assets": ("roa", 200.0),
}
_SECTOR_METRICS = {
    "Sector PE TTM": ("pe", 300.0),
    "Sector Price to Book TTM": ("pb", 60.0),
    "Sector PEG TTM": ("peg", 25.0),
    "Sector Return on Equity ROE": ("roe", 200.0),
    "Sector Return on Assets": ("roa", 200.0),
}
# Company-level columns used to derive whole-market medians.
_COMPANY_METRICS = {
    "PE TTM Price to Earnings": ("pe", 300.0),
    "Price to Book Value Adjusted": ("pb", 60.0),
    "PEG TTM PE to Growth": ("peg", 25.0),
    "ROE Annual %": ("roe", 200.0),
    "RoA Annual %": ("roa", 200.0),
}


# --------------------------------------------------------------------------- io

def _load(path: Path) -> tuple[list[str], list[tuple]]:
    """Read the first worksheet. Vendor files repeat header labels (three
    'Current Price' columns, for example), so callers index by first match."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        book.close()
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return header, rows[1:]


def _index(header: list[str], label: str) -> Optional[int]:
    for i, h in enumerate(header):
        if h.lower() == label.lower():
            return i
    return None


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _number(value: Any, *, cap: Optional[float] = None, zero_is_null: bool = False) -> Optional[float]:
    """Coerce a vendor cell to a float, rejecting sentinels and outliers."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in _TRENDLYNE_NULLS:
            return None
        text = text.replace(",", "").replace("%", "").strip()
        try:
            value = float(text)
        except ValueError:
            return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if out != out or out in (float("inf"), float("-inf")):
        return None
    if zero_is_null and out == 0.0:
        return None
    if cap is not None and abs(out) > cap:
        return None
    return out


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in _TRENDLYNE_NULLS:
        return None
    return text


def file_hash(path: Path) -> Optional[str]:
    if not path.exists():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()[:16]


def _as_of_from_name(path: Path, fallback: str) -> str:
    """Trendlyne names its exports ``YYYY-MM-DD-<group>.xlsx``."""
    stem = path.stem
    parts = stem.split("-")
    if len(parts) >= 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2])).isoformat()
        except (ValueError, TypeError):
            pass
    return fallback


def discover(pattern: str) -> list[Path]:
    return sorted(p for p in _ROOT.glob(pattern) if p.is_file())


def _provenance(source: str, source_version: str) -> dict[str, Any]:
    return {
        "source": source,
        "source_version": source_version,
        "code_version": CODE_VERSION,
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }


# ------------------------------------------------------------- trendlyne parse

def parse_trendlyne_multigroup(path: Path) -> dict[str, Any]:
    """Extract industry medians, sector medians, market medians and ownership.

    Vendor aggregates are repeated on every company row, so industry and sector
    medians are de-duplicated by taking the modal published value per group and
    counting the constituents observed in the file.
    """
    header, body = _load(path)
    if not header:
        return {"ok": False, "error": "empty_workbook", "path": path.name}

    as_of = _as_of_from_name(path, date.today().isoformat())
    version = f"{path.name}:{file_hash(path)}"
    prov = _provenance(SOURCE_TRENDLYNE, version)

    i_sym = _index(header, "NSE Code")
    i_isin = _index(header, "ISIN")
    i_ind = _index(header, "Industry Name")
    i_sec = _index(header, "sector_name")

    industry_acc: dict[tuple[str, str], list[float]] = {}
    industry_members: dict[str, set[str]] = {}
    sector_acc: dict[tuple[str, str], list[float]] = {}
    sector_members: dict[str, set[str]] = {}
    market_acc: dict[str, list[float]] = {}
    ownership_rows: list[dict[str, Any]] = []
    skipped_no_symbol = 0

    own_cols = {
        "promoter_holding": _index(header, "Promoter holding latest %"),
        "institutional_holding": _index(header, "Institutional holding current Qtr %"),
        "fii": _index(header, "FII holding current Qtr %"),
        "mutual_funds": _index(header, "MF holding current Qtr %"),
    }
    i_pledge = _index(header, "Promoter holding pledge percentage % Qtr")
    i_result = _index(header, "Result Announced Date")

    context_rows: list[dict[str, Any]] = []
    ind_idx = {lbl: _index(header, lbl) for lbl in _INDUSTRY_METRICS}
    sec_idx = {lbl: _index(header, lbl) for lbl in _SECTOR_METRICS}
    com_idx = {lbl: _index(header, lbl) for lbl in _COMPANY_METRICS}

    for row in body:
        symbol = _text(_cell(row, i_sym))
        isin = _text(_cell(row, i_isin))
        industry = _text(_cell(row, i_ind))
        sector = _text(_cell(row, i_sec))

        if industry:
            industry_members.setdefault(industry, set()).add(symbol or isin or "?")
            for label, (metric, cap) in _INDUSTRY_METRICS.items():
                v = _number(_cell(row, ind_idx.get(label)), cap=cap)
                if v is not None:
                    industry_acc.setdefault((industry, metric), []).append(v)
        if sector:
            sector_members.setdefault(sector, set()).add(symbol or isin or "?")
            for label, (metric, cap) in _SECTOR_METRICS.items():
                v = _number(_cell(row, sec_idx.get(label)), cap=cap)
                if v is not None:
                    sector_acc.setdefault((sector, metric), []).append(v)

        for label, (metric, cap) in _COMPANY_METRICS.items():
            v = _number(_cell(row, com_idx.get(label)), cap=cap)
            if v is not None:
                market_acc.setdefault(metric, []).append(v)

        # Trendlyne's taxonomy does not match the engine's GICS-style names
        # (BIRLACABLE is "Telecom Cables" here, "Communications Equipment"
        # there). The vendor prints each company's own industry aggregate on
        # its row, so we attach it per symbol and never join on industry name.
        if symbol:
            ctx = {"symbol": symbol, "isin": isin, "as_of": as_of,
                   "vendor_industry": industry, "vendor_sector": sector}
            for label, (metric, cap) in _INDUSTRY_METRICS.items():
                ctx[f"industry_{metric}"] = _number(_cell(row, ind_idx.get(label)), cap=cap)
            for label, (metric, cap) in _SECTOR_METRICS.items():
                ctx[f"sector_{metric}"] = _number(_cell(row, sec_idx.get(label)), cap=cap)
            for label, (metric, cap) in _COMPANY_METRICS.items():
                ctx[f"company_{metric}"] = _number(_cell(row, com_idx.get(label)), cap=cap)
            if any(v is not None for k, v in ctx.items()
                   if k.startswith(("industry_", "sector_", "company_"))):
                context_rows.append({**ctx, **prov})

        # Ownership keys on symbol; ISIN-only rows cannot join company_master.
        if not symbol:
            skipped_no_symbol += 1
            continue
        holdings = {k: _number(_cell(row, i)) for k, i in own_cols.items()}
        if all(v is None for v in holdings.values()):
            continue
        pledge = _number(_cell(row, i_pledge))
        notes = [f"pledge {pledge:.2f}%"] if pledge is not None else []
        if _text(_cell(row, i_result)):
            notes.append(f"result {_text(_cell(row, i_result))}")
        ownership_rows.append({
            "symbol": symbol,
            "as_of": as_of,
            **holdings,
            "confidence": 90,
            "dqiv_status": "VENDOR_REPORTED",
            "validation_notes": "; ".join(notes) or None,
            **prov,
        })

    def _medians(acc, members, key_name):
        out = []
        for (group, metric), values in sorted(acc.items()):
            if not values:
                continue
            out.append({
                key_name: group,
                "metric": metric,
                "as_of": as_of,
                # Vendor repeats one published aggregate per row; median of the
                # repeats recovers it and is robust to stray rows.
                "median_value": round(statistics.median(values), 6),
                "company_count": len(members.get(group, ())) or len(values),
                **prov,
            })
        return out

    market_rows = []
    for metric, values in sorted(market_acc.items()):
        if len(values) < 30:
            continue
        market_rows.append({
            "market": "NSE",
            "metric": metric,
            "as_of": as_of,
            "median_value": round(statistics.median(values), 6),
            "company_count": len(values),
            **prov,
        })

    return {
        "ok": True,
        "path": path.name,
        "as_of": as_of,
        "source_version": version,
        "rows_read": len(body),
        "skipped_no_symbol": skipped_no_symbol,
        "historical_industry_medians": _medians(industry_acc, industry_members, "industry"),
        "historical_sector_medians": _medians(sector_acc, sector_members, "sector"),
        "historical_market_medians": market_rows,
        "ownership": ownership_rows,
        "industry_context": context_rows,
    }


def parse_trendlyne_prices(path: Path) -> dict[str, Any]:
    """Price, range and traded-volume history from the price multigroup export."""
    header, body = _load(path)
    if not header:
        return {"ok": False, "error": "empty_workbook", "path": path.name}

    as_of = _as_of_from_name(path, date.today().isoformat())
    version = f"{path.name}:{file_hash(path)}"

    cols = {
        "symbol": _index(header, "NSE Code"),
        "isin": _index(header, "ISIN"),
        "vwap_day": _index(header, "VWAP Day"),
        "high_1y": _index(header, "1Yr High"),
        "low_1y": _index(header, "1Yr Low"),
        "high_5y": _index(header, "5Yr High"),
        "low_5y": _index(header, "5Yr Low"),
        "high_10y": _index(header, "10Yr High"),
        "low_10y": _index(header, "10Yr Low"),
        "return_1d": _index(header, "Day change %"),
        "return_1m": _index(header, "Month Change %"),
        "return_3m": _index(header, "Qtr Change %"),
        "return_1y": _index(header, "1Yr change %"),
        "return_2y": _index(header, "2Yr price change %"),
        "return_3y": _index(header, "3Yr price change %"),
        "return_5y": _index(header, "5Yr price change %"),
        "return_10y": _index(header, "10Yr price change %"),
        "volume_day": _index(header, "Day Volume"),
        "volume_week_avg": _index(header, "Week Volume Avg"),
        "volume_month_avg": _index(header, "Month Volume Avg"),
        "market_cap": _index(header, "Market Capitalization"),
    }

    out = []
    for row in body:
        symbol = _text(_cell(row, cols["symbol"]))
        if not symbol:
            continue
        rec = {"symbol": symbol, "as_of": as_of, "isin": _text(_cell(row, cols["isin"]))}
        for key, idx in cols.items():
            if key in ("symbol", "isin"):
                continue
            rec[key] = _number(_cell(row, idx))
        out.append(rec)
    return {"ok": True, "path": path.name, "as_of": as_of, "source_version": version,
            "rows_read": len(body), "price_history": out}


def parse_trendlyne_technical(path: Path) -> dict[str, Any]:
    """Beta and technical state. Only present in the older 500-row export —
    the 3,486-row multigroup run omitted this column group."""
    header, body = _load(path)
    if not header:
        return {"ok": False, "error": "empty_workbook", "path": path.name}
    as_of = _as_of_from_name(path, date.today().isoformat())
    version = f"{path.name}:{file_hash(path)}"
    labels = {
        "beta_1m": "Beta 1Month", "beta_3m": "Beta 3Month",
        "beta_1y": "Beta 1Year", "beta_3y": "Beta 3Year",
        "atr": "Day ATR", "adx": "Day ADX", "rsi": "Day RSI", "mfi": "Day MFI",
        "macd": "Day MACD", "macd_signal": "Day MACD Signal Line",
        "sma50": "Day SMA50", "sma200": "Day SMA200",
        "ema20": "Day EMA20", "ema50": "Day EMA50",
        "roc21": "Day ROC21", "roc125": "Day ROC125",
        "momentum_score": "Trendlyne Momentum Score",
        "momentum_score_norm": "Normalized Momentum Score",
        "momentum_score_prev_month": "Prev Month Trendlyne Momentum Score",
        "pivot": "Standard Pivot point",
        "isin": "ISIN",
    }
    cols = {"symbol": _index(header, "NSE Code")}
    cols.update({k: _index(header, v) for k, v in labels.items()})
    out = []
    for row in body:
        symbol = _text(_cell(row, cols["symbol"]))
        if not symbol:
            continue
        rec = {"symbol": symbol, "as_of": as_of, "isin": _text(_cell(row, cols.get("isin")))}
        for key, idx in cols.items():
            if key in ("symbol", "isin"):
                continue
            rec[key] = _number(_cell(row, idx))
        if any(rec[k] is not None for k in rec if k not in ("symbol", "as_of")):
            out.append(rec)
    return {"ok": True, "path": path.name, "as_of": as_of, "source_version": version,
            "rows_read": len(body), "risk_metrics": out}


# ----------------------------------------------------------------- capital iq

def parse_capiq_broker(path: Path) -> dict[str, Any]:
    """Liquidity from the broker export.

    Consensus itself is skipped: 910 of 3,028 rows carry non-zero coverage and
    the warehouse already holds exactly 910 consensus rows, so that side is
    fully ingested. Capital IQ writes 0 rather than blank for "no data", so a
    zero must be read as missing for ratings, targets and volume.
    """
    header, body = _load(path)
    if not header:
        return {"ok": False, "error": "empty_workbook", "path": path.name}

    as_of = date.today().isoformat()
    version = f"{path.name}:{file_hash(path)}"

    i_tic = _index(header, "Ticker")
    i_adv = _index(header, "Daily Volume (Average - 3 Months) [Latest]")
    i_cov = _index(header, "Coverage")
    i_idx = _index(header, "Index Constituents")

    liquidity, covered = [], 0
    for row in body:
        ticker = _text(_cell(row, i_tic))
        if not ticker:
            continue
        if _number(_cell(row, i_cov), zero_is_null=True) is not None:
            covered += 1
        adv = _number(_cell(row, i_adv), zero_is_null=True)
        if adv is None:
            continue
        membership = _text(_cell(row, i_idx))
        liquidity.append({
            "symbol": ticker,
            "as_of": as_of,
            "adv_3m": adv,
            # Current membership only — this is not point-in-time and cannot be
            # used to reconstruct a survivorship-free historical universe.
            "index_membership": (membership.replace("\n\n", "; ") if membership and membership != "0" else None),
        })
    return {"ok": True, "path": path.name, "as_of": as_of, "source_version": version,
            "rows_read": len(body), "consensus_covered": covered, "liquidity": liquidity}


# --------------------------------------------------------------------- driver

def collect(*, root: Path = _ROOT) -> dict[str, Any]:
    """Parse every checked-in vendor export and return warehouse-ready rows."""
    bundle: dict[str, list[dict[str, Any]]] = {}
    reports: list[dict[str, Any]] = []

    def merge(result: dict[str, Any], keys: Iterable[str]) -> None:
        reports.append({k: v for k, v in result.items()
                        if k not in ("historical_industry_medians", "historical_sector_medians",
                                     "historical_market_medians", "ownership", "price_history",
                                     "industry_context",
                                     "risk_metrics", "liquidity")})
        for key in keys:
            if result.get(key):
                bundle.setdefault(key, []).extend(result[key])

    for path in discover("*multigroup*.xlsx"):
        header, _ = _load(path)
        if _index(header, "Promoter holding latest %") is not None:
            merge(parse_trendlyne_multigroup(path),
                  ("historical_industry_medians", "historical_sector_medians",
                   "historical_market_medians", "ownership", "industry_context"))
        elif _index(header, "10Yr High") is not None:
            merge(parse_trendlyne_prices(path), ("price_history",))
        elif _index(header, "Beta 1Year") is not None:
            merge(parse_trendlyne_technical(path), ("risk_metrics",))

    tech = root / "trendlyne_technical_ownership.xlsx"
    if tech.exists():
        legacy = parse_trendlyne_technical(tech)
        seen = {r["symbol"] for r in bundle.get("risk_metrics", ())}
        legacy["risk_metrics"] = [r for r in legacy.get("risk_metrics", ())
                                  if r["symbol"] not in seen]
        merge(legacy, ("risk_metrics",))

    broker = root / "capital_iq_exports" / "broker_estimates.xlsx"
    if broker.exists():
        merge(parse_capiq_broker(broker), ("liquidity",))

    return {"ok": True, "reports": reports,
            "counts": {k: len(v) for k, v in sorted(bundle.items())},
            "rows": bundle}


# Tabs that exist in institutional_warehouse.schema and can be written directly.
WAREHOUSE_TABS = ("historical_industry_medians", "historical_sector_medians",
                  "historical_market_medians", "ownership")


def write(bundle: dict[str, list[dict[str, Any]]], *, actor: str = "vendor_exports") -> dict[str, Any]:
    """Persist the tabs that already exist in the warehouse schema.

    ``price_history``, ``risk_metrics`` and ``liquidity`` have no tab yet and
    are returned for a caller to route once one is defined, rather than being
    silently dropped.
    """
    from institutional_warehouse import gateway

    written: dict[str, Any] = {}
    for tab in WAREHOUSE_TABS:
        rows = bundle.get(tab) or []
        if not rows:
            continue
        source = rows[0].get("source", SOURCE_TRENDLYNE)
        gateway.write(tab, rows, source=source, actor=actor)
        written[tab] = len(rows)

    unrouted = {k: len(v) for k, v in bundle.items() if k not in WAREHOUSE_TABS and v}
    return {"ok": True, "written": written, "unrouted": unrouted}
