"""Capital IQ estimate vintages — what consensus believed on each past date.

This is the dataset that separates an honest backtest from a flattered one.
A normal estimate pull answers "what is the FY1 EPS estimate", meaning today's
consensus. A vintage answers "what did the market believe on 2021-03-31".
Ranking a 2021 portfolio on today's consensus is look-ahead bias: analysts had
already revised toward the outcome.

Source workbook (capiq_vintage_template.xlsx, pulled 2026-08-19):

    EPS EST     72 monthly columns, 2020-01 to 2025-12
                1,099 of 3,023 companies carry data (36.4%)
    Basic EPS   same grid, reported EPS as known at each date
                3,006 of 3,023 companies (99.4%)

Estimate coverage is not a defect. Only ~900-1,100 Indian listings carry
analyst coverage at all, and the count grows across the window (648 in 2020 to
801 in 2025), which is the market rather than the export.

Two conventions this module encodes:

* Capital IQ writes 0 for "no data", the same behaviour seen in
  broker_estimates.xlsx. A zero here means no analyst covered the name that
  month, not an EPS forecast of zero. Treating it as data would put thousands
  of fictitious zeros into every factor built on this.
* The workbook's period-end column returned "(Invalid Time Period)" for every
  row, so the fiscal year each FY1 refers to is derived from the as-of date
  instead. Indian fiscal years end 31 March, so an as-of date in Jan-Mar has
  FY1 ending that same March, and Apr-Dec has FY1 ending the following March.
  Every row records period_source so a derived label is never mistaken for a
  vendor-supplied one.
"""

from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Optional

_ROOT = Path(__file__).resolve().parents[2]

SOURCE = "capital_iq_estimate_vintages"
CODE_VERSION = "capiq_vintages_v1"
DEFAULT_WORKBOOK = _ROOT / "capiq_vintage_template.xlsx"

# Sheet -> (metric name, whether the figure is a forward estimate)
SHEETS: dict[str, tuple[str, bool]] = {
    "EPS EST": ("eps_estimate", True),
    "Basic EPS": ("eps_reported", False),
    # Tolerate the original template names too.
    "Vintages_EPS": ("eps_estimate", True),
    "Vintages_NumEst": ("analyst_count", True),
}

_ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}\d$")


def _clean_isin(value: Any) -> Optional[str]:
    """Capital IQ prefixes ISINs with I_ (e.g. I_INE144J01027)."""
    text = str(value or "").strip().upper()
    if text.startswith("I_"):
        text = text[2:]
    return text if _ISIN_RE.match(text) else None


def _clean_symbol(value: Any) -> Optional[str]:
    """Identifiers arrive as EXCHANGE:TICKER (e.g. NSEI:20MICRONS)."""
    text = str(value or "").strip().upper()
    if ":" in text:
        text = text.split(":", 1)[1]
    text = text.strip()
    return text or None


def _number(value: Any) -> Optional[float]:
    """Coerce a cell, treating Capital IQ's 0 sentinel as missing.

    A genuine reported EPS of exactly 0.0 is vanishingly rare and
    indistinguishable from the sentinel, so both are dropped. Losing a true
    zero is far cheaper than admitting thousands of false ones.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("("):  # "(Invalid Time Period)"
            return None
        try:
            value = float(text.replace(",", ""))
        except ValueError:
            return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if out != out or out in (float("inf"), float("-inf")) or out == 0.0:
        return None
    return out


def fiscal_period(as_of: date, *, forward: bool) -> tuple[str, date]:
    """Which fiscal year an FY1 figure refers to, for an Indian fiscal calendar.

    FY ends 31 March. An as-of date in Jan-Mar sits inside the FY ending that
    March; Apr-Dec sits inside the FY ending the following March. A forward
    estimate points at the FY the date sits in; a reported figure refers to the
    most recently completed one.
    """
    fy_end_year = as_of.year if as_of.month <= 3 else as_of.year + 1
    if not forward:
        fy_end_year -= 1
    return f"FY{fy_end_year}", date(fy_end_year, 3, 31)


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def file_hash(path: Path) -> Optional[str]:
    if not path.exists():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()[:16]


def _iter_sheet(path: Path, sheet: str) -> Iterator[tuple[list, list]]:
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            return
        ws = book[sheet]
        header: Optional[list] = None
        for row in ws.iter_rows(values_only=True):
            # The grid header is the row whose first cell is exactly "ISIN".
            if header is None:
                if row and str(row[0] or "").strip().upper() == "ISIN":
                    header = list(row)
                continue
            yield header, list(row)
    finally:
        book.close()


def parse(path: Path = DEFAULT_WORKBOOK) -> dict[str, Any]:
    """Long-format vintage rows, one per (symbol, as-of date, metric)."""
    if not path.exists():
        return {"ok": False, "error": f"workbook_not_found:{path.name}", "rows": []}

    version = f"{path.name}:{file_hash(path)}"
    rows: list[dict[str, Any]] = []
    stats: dict[str, dict[str, int]] = {}

    for sheet, (metric, forward) in SHEETS.items():
        seen = {"companies": 0, "with_data": 0, "cells": 0, "zeros_dropped": 0,
                "no_symbol": 0}
        dates: list[Optional[date]] = []
        for header, raw in _iter_sheet(path, sheet):
            if not dates:
                dates = [_as_date(h) for h in header[3:]]
            seen["companies"] += 1
            isin = _clean_isin(raw[0] if len(raw) > 0 else None)
            symbol = _clean_symbol(raw[1] if len(raw) > 1 else None)
            if not symbol:
                seen["no_symbol"] += 1
                continue
            wrote = 0
            for offset, as_of in enumerate(dates):
                if as_of is None:
                    continue
                idx = 3 + offset
                if idx >= len(raw):
                    break
                value = _number(raw[idx])
                if value is None:
                    if raw[idx] == 0:
                        seen["zeros_dropped"] += 1
                    continue
                period, period_end = fiscal_period(as_of, forward=forward)
                rows.append({
                    "symbol": symbol,
                    "isin": isin,
                    "consensus_date": as_of.isoformat(),
                    "target_period": period,
                    "target_period_end": period_end.isoformat(),
                    # Derived, not vendor-supplied — the workbook's period-end
                    # column returned "(Invalid Time Period)" throughout.
                    "period_source": "derived_indian_fy",
                    "metric": metric,
                    "mean_estimate": round(value, 6),
                    "currency": "INR",
                    "unit": "per_share",
                    "is_forward_estimate": "true" if forward else "false",
                    "source": SOURCE,
                    "source_version": version,
                    "code_version": CODE_VERSION,
                })
                wrote += 1
            seen["cells"] += wrote
            if wrote:
                seen["with_data"] += 1
        if seen["companies"]:
            stats[sheet] = seen

    return {"ok": True, "path": path.name, "source_version": version,
            "rows": rows, "stats": stats}


def summarise(parsed: dict[str, Any]) -> dict[str, Any]:
    """Coverage summary — what a consumer needs before trusting a backtest."""
    rows = parsed.get("rows") or []
    if not rows:
        return {"ok": False, "rows": 0}
    by_metric: dict[str, set] = {}
    by_year: dict[int, set] = {}
    dates = set()
    for r in rows:
        by_metric.setdefault(r["metric"], set()).add(r["symbol"])
        year = int(r["consensus_date"][:4])
        by_year.setdefault(year, set()).add(r["symbol"])
        dates.add(r["consensus_date"])
    return {
        "ok": True,
        "rows": len(rows),
        "symbols": len({r["symbol"] for r in rows}),
        "date_range": [min(dates), max(dates)],
        "months": len(dates),
        "symbols_by_metric": {k: len(v) for k, v in sorted(by_metric.items())},
        "symbols_by_year": {k: len(v) for k, v in sorted(by_year.items())},
    }


# The warehouse tab keys on (symbol, consensus_date, target_period, metric),
# which is exactly this grain. Reported figures share the tab: `metric`
# distinguishes eps_reported from eps_estimate, and is_forward_estimate makes
# the distinction explicit for anything that reads the rows directly.
WAREHOUSE_TAB = "consensus_metric_vintages"


def write(parsed: dict[str, Any], *, actor: str = "capiq_vintages") -> dict[str, Any]:
    from institutional_warehouse import gateway

    rows = parsed.get("rows") or []
    if not rows:
        return {"ok": False, "written": 0, "error": "no_rows"}
    gateway.write(WAREHOUSE_TAB, rows, source=SOURCE, actor=actor,
                  reason="capiq_estimate_vintage_import")
    return {"ok": True, "written": len(rows), "tab": WAREHOUSE_TAB}
