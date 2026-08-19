"""Forward EPS and revenue estimates, the current cross-section.

`capiq_vintages` reads the historical vintage sheets of the same workbook - what
was believed on each month end from 2020 to 2025. Those are the right input for
a point-in-time backtest and the wrong one for a live screen, which needs what
is believed *now*. The `Forward_Estimates_Now` sheet carries that, and nothing
read it, so `forward_pe` was null across the universe and two desk screens -
Forward Earnings Growth and Alpha Opportunity - returned zero rows.

Coverage is 910 companies of 3,023. That is not an export failure: it matches
the `consensus` tab's 910 exactly, and it is simply how many Indian listed
companies carry sell-side coverage in Capital IQ. The remaining ~2,100 are
uncovered micro caps with genuinely no estimate, so a screen that requires a
forward number will always be a minority of the universe and should say so
rather than look broken.

Two conventions in this export would corrupt the result silently:

* Capital IQ writes 0 for "no data". Roughly 70% of the estimate cells are that
  sentinel, so reading them as forecasts would put a zero EPS - an infinite
  forward P/E - against two thousand companies.
* The `Period end FY1` column is empty for every row, the same
  "(Invalid Time Period)" failure the vintage sheets hit. The fiscal period is
  therefore derived from the Indian fiscal calendar and labelled as derived.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Iterable, Optional

from .capiq_vintages import _clean_isin, _clean_symbol, _number, fiscal_period

SOURCE = "capital_iq_forward_estimates"
SHEET = "Forward_Estimates_Now"
WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "capiq_vintage_template.xlsx"


def default_as_of() -> date:
    """The date FY1 is measured from: the import date.

    This sheet is a live pull, so "FY1" means whatever Capital IQ considered
    the next unreported year when the workbook was last refreshed. Dating it
    from the workbook's *vintage* columns instead - they end 2025-12-31 - was
    wrong twice over: it labelled the estimates FY2026, a year that closed in
    March, and it collided with the genuine 2025-12-31 vintage rows on
    (symbol, consensus_date, target_period, metric), overwriting 837 of them
    with figures for a different year.

    Today is the right anchor as long as the workbook is imported near when it
    was pulled. If it goes stale the label drifts, so the import reports the
    as_of it used and flags a period that has already closed.
    """
    return date.today()

# (column index, metric, target period offset in years from FY1)
_FIELDS = (
    (2, "eps_estimate", 0),
    (3, "eps_estimate", 1),
    (4, "revenue_estimate", 0),
    (5, "revenue_estimate", 1),
)


def _rows(path: Path, as_of: date) -> list[dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[SHEET]
        label_fy1, end_fy1 = fiscal_period(as_of, forward=True)
        out: list[dict[str, Any]] = []
        for raw in sheet.iter_rows(min_row=4, values_only=True):
            if not raw or len(raw) < 6:
                continue
            isin = _clean_isin(raw[0])
            symbol = _clean_symbol(raw[1])
            if not isin or not symbol:
                continue
            for index, metric, offset in _FIELDS:
                value = _number(raw[index] if index < len(raw) else None)
                if value is None:
                    continue
                if offset == 0:
                    label, period_end = label_fy1, end_fy1
                else:
                    label = f"FY{int(label_fy1[2:]) + offset}"
                    period_end = date(end_fy1.year + offset, end_fy1.month, end_fy1.day)
                out.append({
                    "symbol": symbol,
                    "isin": isin,
                    "consensus_date": as_of.isoformat(),
                    "target_period": label,
                    "target_period_end": period_end.isoformat(),
                    "metric": metric,
                    "mean_estimate": value,
                    "currency": "INR",
                    # Derived, never vendor-supplied: the export's own period
                    # column is empty for every row.
                    "period_source": "derived_indian_fy",
                    "is_forward_estimate": "true",
                    "source": SOURCE,
                })
        return out
    finally:
        workbook.close()


def parse(*, path: Path = WORKBOOK_PATH, as_of: Optional[date] = None) -> dict[str, Any]:
    as_of = as_of or default_as_of()
    if not path.exists():
        return {"ok": False, "error": f"workbook_not_found:{path.name}", "rows": []}
    rows = _rows(path, as_of)
    label_fy1, end_fy1 = fiscal_period(as_of, forward=True)
    stale = end_fy1 < date.today()
    symbols = {r["symbol"] for r in rows}
    by_metric: dict[str, int] = {}
    for row in rows:
        by_metric[row["metric"]] = by_metric.get(row["metric"], 0) + 1
    return {
        "ok": bool(rows),
        "source": SOURCE,
        "workbook": path.name,
        "sheet": SHEET,
        "as_of": as_of.isoformat(),
        "rows": rows,
        "row_count": len(rows),
        "symbols": len(symbols),
        "by_metric": by_metric,
        "stale": stale,
        "limitations": ([
            f"STALE: FY1 resolves to {label_fy1}, which ended {end_fy1.isoformat()} - "
            "already closed. The workbook needs refreshing; these are not forward "
            "estimates any more."
        ] if stale else []) + [
            "Capital IQ writes 0 for no-data; those cells are dropped rather than "
            "read as a zero forecast, which would imply an infinite forward P/E.",
            "Fiscal periods are derived from the Indian fiscal calendar because the "
            "export's own period column is empty for every row.",
            "Coverage is roughly 910 of 3,023 companies - the real extent of "
            "sell-side coverage, not a truncated export.",
        ],
    }


def import_estimates(*, actor: str = "fwcp", path: Path = WORKBOOK_PATH,
                     as_of: Optional[date] = None) -> dict[str, Any]:
    """Write the cross-section into consensus_metric_vintages."""
    from institutional_warehouse import gateway

    parsed = parse(path=path, as_of=as_of)
    if not parsed.get("ok"):
        return parsed
    written = gateway.write(
        "consensus_metric_vintages", parsed["rows"], source=SOURCE, actor=actor,
        reason="capiq_forward_estimates:current_cross_section",
    )
    return {**{k: v for k, v in parsed.items() if k != "rows"}, "written": written,
            "ok": bool(written.get("ok"))}


def latest_forward_eps(rows: Optional[Iterable[dict[str, Any]]] = None) -> dict[str, float]:
    """{symbol: FY1 EPS} for deriving a forward P/E.

    Reads the warehouse when given nothing, so the scanner does not have to
    know where the estimates came from.
    """
    if rows is None:
        try:
            from institutional_warehouse import store

            rows = store.all_rows("consensus_metric_vintages", limit=400000) or []
        except Exception:
            return {}
    # Newest consensus date wins; within it the nearest target period is FY1.
    # FY2 must never be substituted - it would understate the multiple and make
    # a company look cheaper on forward earnings than anyone actually forecast.
    best: dict[str, tuple[str, str, float]] = {}
    for row in rows or []:
        if str(row.get("metric") or "") != "eps_estimate":
            continue
        if str(row.get("is_forward_estimate") or "").lower() != "true":
            continue
        symbol = str(row.get("symbol") or "").upper()
        stamp = str(row.get("consensus_date") or "")
        period = str(row.get("target_period") or "")
        value = _number(row.get("mean_estimate"))
        if not symbol or not stamp or not period or value is None or value <= 0:
            continue
        prev = best.get(symbol)
        if prev is None:
            best[symbol] = (stamp, period, value)
            continue
        newer_date = stamp > prev[0]
        nearer_period = stamp == prev[0] and period < prev[1]
        if newer_date or nearer_period:
            best[symbol] = (stamp, period, value)
    return {symbol: value for symbol, (_, _, value) in best.items()}
