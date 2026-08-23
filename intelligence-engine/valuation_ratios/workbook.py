"""Admin workbook: one sheet per ratio, companies down column A, dates across.

The warehouse stores ``valuation_ratios`` long - one row per (symbol,
ratio_name, reported_date, snapshot_id). That shape answers "what is
RELIANCE's P/E today" and is wrong for the question the desk actually asks,
which is "show me every company's P/E and how it has moved". Pivoting 62k+
rows by hand in Excel is the sort of thing that gets done once and then
silently goes stale, so it is built here instead.

Sheets are the six ratios Upstox reports, plus a coverage sheet. Six ratios
is not a design choice - it is what ``/v2/fundamentals/{isin}/key-ratios``
returns, and inventing a seventh would mean inventing the data behind it.

Dates run newest-first from column D. A year of daily collection is 250
columns; appending on the right would mean scrolling to the far edge of the
sheet to see today. The sheet is rebuilt from the warehouse on each request
rather than incrementally appended, because the warehouse is the record and a
workbook that drifts from it is worse than no workbook.

Every company in the eligible universe gets a row even when it has no data.
An admin sheet exists to show gaps; dropping the empty rows would hide exactly
the companies worth looking at.

Both the workbook and the pivot are built for a process that has run out of
memory before. openpyxl's normal mode held 940 MB for 2,400 companies by 250
dates and write-only holds 51 MB for the same sheet, so this streams. The
pivot is per-ratio arrays aligned to the date list rather than one dict keyed
by (symbol, ratio, date) - the tuple-keyed version measured 403 MB where the
arrays cost a fraction of it.
"""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Optional

from valuation_ratios.sweep import ELIGIBLE_EQUITY, EXPECTED, classify

# Display names. The tab label is what an admin reads, so "P/E" rather than
# "pe" - but "/" is illegal in an Excel sheet name, hence the spelled forms.
SHEET_TITLES: dict[str, str] = {
    "pe": "P-E",
    "pb": "P-B",
    "roa": "ROA",
    "roe": "ROE",
    "roce": "ROCE",
    "ev_ebitda": "EV-EBITDA",
}

COVERAGE_SHEET = "Coverage"
IDENTITY_HEADERS = ("Symbol", "Company", "Sector")
DEFAULT_DAYS = 120
# About two trading years. Past this the pivot and the response both grow
# faster than they are worth: 750 dates is a 28 MB download and roughly forty
# seconds of building, which is long enough for a proxy to give up on it.
MAX_DAYS = 500


def _physical() -> str:
    from institutional_warehouse import db

    return db.physical_table("valuation_ratios")


def _universe() -> list[dict[str, Any]]:
    """Eligible equities from the master, plus anything that already has rows.

    The second half matters: a symbol can carry ratios collected before the
    classifier learned to read it, and dropping it here would quietly shrink
    the sheet relative to the warehouse it is supposed to mirror.
    """
    from institutional_warehouse import db, store

    known: dict[str, dict[str, Any]] = {}
    for row in store.all_rows("company_master", limit=20000) or []:
        symbol = str(row.get("symbol") or "").strip().upper()
        if not symbol:
            continue
        known[symbol] = {
            "symbol": symbol,
            "company": str(row.get("company_name") or "").strip(),
            "sector": str(row.get("sector") or "").strip(),
            "isin": str(row.get("isin") or "").strip().upper(),
            "eligible": classify(row) == ELIGIBLE_EQUITY,
        }

    with_rows = {
        str(r.get("symbol") or "").strip().upper()
        for r in db.query(f"SELECT DISTINCT symbol FROM {_physical()}")
    }
    out = [c for c in known.values() if c["eligible"] or c["symbol"] in with_rows]
    for symbol in sorted(with_rows - set(known)):
        if symbol:
            out.append({"symbol": symbol, "company": "", "sector": "",
                        "isin": "", "eligible": False})
    out.sort(key=lambda c: c["symbol"])
    return out


def recent_dates(days: int) -> list[str]:
    """The most recent N collection dates, newest first.

    Dates come from the data rather than from a calendar. A market holiday is
    not a missing column, and back-dating a column the sweep never ran would
    turn "we did not collect" into "we collected nothing", which are different
    facts.
    """
    from institutional_warehouse import db

    rows = db.query(
        f"SELECT DISTINCT reported_date AS d FROM {_physical()} "
        "WHERE reported_date IS NOT NULL AND reported_date <> '' "
        "ORDER BY d DESC"
    )
    dates = [str(r["d"]) for r in rows if str(r.get("d") or "").strip()]
    return dates[:days]


def series_by_ratio(dates: list[str]) -> dict[str, dict[str, list[Optional[float]]]]:
    """{ratio: {symbol: [value per date, aligned to ``dates``]}}.

    Arrays rather than a dict keyed by (symbol, ratio, date): the tuple-keyed
    form measured 403 MB at 3.6 million entries, which this process cannot
    afford alongside everything else it runs.

    The natural key carries ``snapshot_id``, so a day re-swept holds more than
    one row per company and ratio. Ordering by reported_time then snapshot_id
    and letting the last write win means the sheet shows the most recent
    reading rather than whichever row the database happened to return first.
    """
    from institutional_warehouse import db

    out: dict[str, dict[str, list[Optional[float]]]] = {r: {} for r in EXPECTED}
    if not dates:
        return out
    index = {day: position for position, day in enumerate(dates)}
    placeholders = ",".join(["?"] * len(dates))
    # "?" on both backends: db._bind rewrites it to named binds for
    # SQLAlchemy, so writing dialect-specific placeholders here would break
    # Postgres rather than support it.
    rows = db.query(
        "SELECT symbol, ratio_name, reported_date, company_value "
        f"FROM {_physical()} WHERE reported_date IN ({placeholders}) "
        "ORDER BY reported_time ASC, snapshot_id ASC",
        tuple(dates),
    )
    for row in rows:
        value = row.get("company_value")
        if value is None:
            continue
        bucket = out.get(str(row.get("ratio_name") or "").strip().lower())
        if bucket is None:
            continue
        position = index.get(str(row.get("reported_date") or ""))
        if position is None:
            continue
        symbol = str(row.get("symbol") or "").strip().upper()
        line = bucket.get(symbol)
        if line is None:
            line = bucket[symbol] = [None] * len(dates)
        try:
            line[position] = float(value)
        except (TypeError, ValueError):
            continue
    return out


def _prepare(sheet, headers: list[str], date_columns: int) -> None:
    """Layout must be set before rows stream out; write-only cannot go back."""
    from openpyxl.utils import get_column_letter

    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1048576"
    for column, width in ((1, 16), (2, 38), (3, 22)):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for offset in range(date_columns):
        sheet.column_dimensions[get_column_letter(4 + offset)].width = 11


def _header_row(sheet, headers: list[str], *, tilt_from: int = 0):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    tilted = Alignment(textRotation=60, horizontal="center")
    cells = []
    for position, title in enumerate(headers, start=1):
        cell = WriteOnlyCell(sheet, value=title)
        cell.font = bold
        if tilt_from and position > tilt_from:
            cell.alignment = tilted
        cells.append(cell)
    return cells


def build_bytes(*, days: int = DEFAULT_DAYS) -> tuple[bytes, dict[str, Any]]:
    """Serialise to memory. The engine has no writable disk it should rely on."""
    import io

    from openpyxl import Workbook

    days = max(1, min(int(days or DEFAULT_DAYS), MAX_DAYS))
    universe = _universe()
    dates = recent_dates(days)
    series = series_by_ratio(dates)
    latest = dates[0] if dates else None

    workbook = Workbook(write_only=True)
    blank: list[Optional[float]] = [None] * len(dates)
    values = 0

    for ratio in EXPECTED:
        sheet = workbook.create_sheet(SHEET_TITLES[ratio])
        headers = list(IDENTITY_HEADERS) + list(dates)
        _prepare(sheet, headers, len(dates))
        sheet.append(_header_row(sheet, headers, tilt_from=len(IDENTITY_HEADERS)))
        bucket = series[ratio]
        for company in universe:
            line = bucket.get(company["symbol"], blank)
            values += sum(1 for v in line if v is not None)
            sheet.append([company["symbol"], company["company"],
                          company["sector"], *line])

    coverage = workbook.create_sheet(COVERAGE_SHEET)
    headers = (
        list(IDENTITY_HEADERS)
        # Names the date rather than saying "latest", because the column
        # beside it is the company's own latest date and the two differ for
        # any company the newest sweep did not reach.
        + ["ISIN", "Eligible", "Days Collected", "First Date", "Latest Date",
           f"Ratios On {latest}" if latest else "Ratios On Newest Date"]
        + [f"Latest {SHEET_TITLES[r]}" for r in EXPECTED]
    )
    _prepare(coverage, headers, 0)
    coverage.append(_header_row(coverage, headers))

    covered = 0
    for company in universe:
        symbol = company["symbol"]
        lines = [series[r].get(symbol, blank) for r in EXPECTED]
        collected = [dates[i] for i in range(len(dates))
                     if any(line[i] is not None for line in lines)]
        on_latest = sum(1 for line in lines if line and line[0] is not None)
        covered += 1 if on_latest else 0
        # First non-empty walking newest-first is the most recent reading,
        # which is not the same as the value on the latest date - a ratio
        # missing today still has a last known value, and blanking it would
        # report a collection gap as an absent ratio.
        latest_each = [next((v for v in line if v is not None), None) for line in lines]
        coverage.append(
            [symbol, company["company"], company["sector"], company["isin"],
             "yes" if company["eligible"] else "no", len(collected),
             collected[-1] if collected else None,
             collected[0] if collected else None, on_latest, *latest_each]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    summary = {
        "companies": len(universe),
        "dates": len(dates),
        "latest_date": latest,
        "oldest_date": dates[-1] if dates else None,
        "values": values,
        "companies_on_latest_date": covered,
        "sheets": [SHEET_TITLES[r] for r in EXPECTED] + [COVERAGE_SHEET],
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "bytes": len(payload),
    }
    return payload, summary


def summarise(*, days: int = DEFAULT_DAYS) -> dict[str, Any]:
    """What the workbook would contain, without holding the file."""
    return build_bytes(days=days)[1]


def filename(summary: Optional[dict[str, Any]] = None) -> str:
    stamp = (summary or {}).get("latest_date") or datetime.now(timezone.utc).date().isoformat()
    return f"agi_valuation_ratios_{stamp}.xlsx"
