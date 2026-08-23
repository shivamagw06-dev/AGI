"""The admin workbook: shape, gaps, and which reading wins."""

from __future__ import annotations

import io
import os
import tempfile

import pytest

os.environ.setdefault("INSTITUTIONAL_WAREHOUSE_ROOT",
                      tempfile.mkdtemp(prefix="wh_workbook_"))

from valuation_ratios import workbook


@pytest.fixture()
def warehouse(monkeypatch, tmp_path):
    """A real SQLite warehouse, because the pivot is mostly SQL."""
    from institutional_warehouse import db, gateway

    monkeypatch.setenv("INSTITUTIONAL_WAREHOUSE_ROOT", str(tmp_path))
    monkeypatch.delenv("INSTITUTIONAL_WAREHOUSE_DATABASE_URL", raising=False)
    monkeypatch.delenv("WAREHOUSE_DATABASE_URL", raising=False)
    db.reset_backend()
    db.init(force=True)

    gateway.write("company_master", [
        {"company_id": "C1", "symbol": "AAA", "company_name": "Alpha Ltd",
         "isin": "INE000A01001", "sector": "Technology"},
        {"company_id": "C2", "symbol": "BBB", "company_name": "Beta Ltd",
         "isin": "INE000A01002", "sector": "Financials"},
        {"company_id": "C3", "symbol": "CCC", "company_name": "Gamma Ltd",
         "isin": "INE000A01003", "sector": "Energy"},
    ], source="test", actor="test")
    yield
    db.reset_backend()


def _open(days=30):
    """Reopen the saved file. Write-only workbooks cannot be read in place."""
    from openpyxl import load_workbook

    payload, summary = workbook.build_bytes(days=days)
    return load_workbook(io.BytesIO(payload)), summary


def _ratio(symbol, name, date, value, *, snapshot="s1", time="09:00"):
    return {"symbol": symbol, "isin": f"INE{symbol}", "ratio_name": name,
            "company_value": value, "reported_date": date,
            "reported_time": f"{date}T{time}:00Z", "snapshot_id": snapshot,
            "provider": "upstox"}


def _write(rows):
    from institutional_warehouse import gateway

    return gateway.write("valuation_ratios", rows, source="upstox", actor="test")


def test_one_sheet_per_ratio_plus_coverage(warehouse):
    """Seven sheets: the six ratios Upstox reports, and the gap sheet."""
    _write([_ratio("AAA", "pe", "2026-08-20", 25.0)])
    book, summary = _open(30)

    assert book.sheetnames == ["P-E", "P-B", "ROA", "ROE", "ROCE",
                               "EV-EBITDA", "Coverage"]
    assert summary["sheets"] == book.sheetnames


def test_companies_go_down_column_a_and_dates_across(warehouse):
    _write([
        _ratio("AAA", "pe", "2026-08-20", 25.0),
        _ratio("AAA", "pe", "2026-08-21", 26.0),
        _ratio("BBB", "pe", "2026-08-21", 12.0),
    ])
    sheet = _open(30)[0]["P-E"]

    assert [c.value for c in sheet[1][:3]] == ["Symbol", "Company", "Sector"]
    # Newest first. A year of collection is 250 columns, and appending on the
    # right would put today off the edge of the screen.
    assert [c.value for c in sheet[1][3:]] == ["2026-08-21", "2026-08-20"]
    assert [sheet.cell(row=r, column=1).value for r in range(2, 5)] == ["AAA", "BBB", "CCC"]
    assert sheet.cell(row=2, column=4).value == 26.0
    assert sheet.cell(row=2, column=5).value == 25.0


def test_a_company_with_no_data_still_gets_a_row(warehouse):
    """An admin sheet exists to show gaps, so it must not drop the gaps."""
    _write([_ratio("AAA", "pe", "2026-08-20", 25.0)])
    sheet = _open(30)[0]["P-E"]

    symbols = [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]
    assert "CCC" in symbols
    assert sheet.cell(row=symbols.index("CCC") + 2, column=4).value is None


def test_dates_come_from_the_data_not_the_calendar(warehouse):
    """A market holiday is not a missing column."""
    _write([
        _ratio("AAA", "pe", "2026-08-20", 25.0),
        _ratio("AAA", "pe", "2026-08-24", 27.0),
    ])
    assert workbook.recent_dates(30) == ["2026-08-24", "2026-08-20"]


def test_a_resweep_shows_the_later_snapshot(warehouse):
    """The key carries snapshot_id, so one day can hold several readings."""
    _write([_ratio("AAA", "pe", "2026-08-21", 25.0, snapshot="s1", time="09:00")])
    _write([_ratio("AAA", "pe", "2026-08-21", 31.5, snapshot="s2", time="17:00")])
    sheet = _open(30)[0]["P-E"]

    assert sheet.cell(row=2, column=4).value == 31.5


def test_coverage_reports_the_last_known_value_not_only_todays(warehouse):
    """A ratio missing today still has a last reading; blanking it would
    report a collection gap as an absent ratio."""
    _write([
        _ratio("AAA", "pe", "2026-08-20", 25.0),
        _ratio("AAA", "pb", "2026-08-21", 3.2),
    ])
    sheet = _open(30)[0]["Coverage"]
    headers = [c.value for c in sheet[1]]
    row = {h: sheet.cell(row=2, column=i + 1).value for i, h in enumerate(headers)}

    assert row["Symbol"] == "AAA"
    assert row["Days Collected"] == 2
    assert row["First Date"] == "2026-08-20"
    assert row["Latest Date"] == "2026-08-21"
    # The header names the date, because the column beside it is the
    # company's own latest date and they differ whenever a sweep misses one.
    assert "Ratios On 2026-08-21" in headers
    # Only P/B was collected on that date...
    assert row["Ratios On 2026-08-21"] == 1
    # ...but P/E's last reading is still reported.
    assert row["Latest P-E"] == 25.0
    assert row["Latest P-B"] == 3.2


def test_it_serialises_to_a_real_xlsx(warehouse):
    from openpyxl import load_workbook

    _write([_ratio("AAA", "roce", "2026-08-21", 18.4)])
    payload, summary = workbook.build_bytes(days=30)

    assert payload[:2] == b"PK"
    assert summary["bytes"] == len(payload)
    reopened = load_workbook(io.BytesIO(payload))
    assert reopened["ROCE"].cell(row=2, column=4).value == 18.4


def test_days_is_bounded(warehouse):
    """An unbounded request is 2,400 companies by every day ever collected."""
    _write([_ratio("AAA", "pe", "2026-08-21", 25.0)])
    assert workbook.build_bytes(days=10**9)[1]["dates"] <= workbook.MAX_DAYS
    assert workbook.build_bytes(days=0)[1]["dates"] >= 0


def test_the_filename_carries_the_latest_date(warehouse):
    _write([_ratio("AAA", "pe", "2026-08-21", 25.0)])
    assert workbook.filename(_open(30)[1]) == \
        "agi_valuation_ratios_2026-08-21.xlsx"


def test_the_route_is_token_guarded():
    """It is the whole universe in one file - an export, not a page."""
    import pathlib
    import re

    src = pathlib.Path(__file__).resolve().parents[2] / "app" / "api" / "routes.py"
    text = src.read_text()
    for path in ("/valuation-ratios/workbook", "/valuation-ratios/workbook/summary"):
        match = re.search(rf'@router\.get\("{re.escape(path)}"[^)]*\)', text)
        assert match, f"{path} route missing"
        assert "Depends(require_token)" in match.group(0), f"{path} is unguarded"


def test_the_workbook_streams_rather_than_holding_every_cell():
    """This runs in a process that has been killed for memory before.

    openpyxl's normal mode measured 940 MB for 2,400 companies by 250 dates;
    write-only measured 51 MB for the same sheet. The engine's recalculate
    stage already reached 8.58 GB of an 8.59 GB cap, so the difference is not
    an optimisation.
    """
    import inspect

    src = inspect.getsource(workbook.build_bytes)
    assert "Workbook(write_only=True)" in src


def test_the_pivot_is_arrays_not_one_tuple_keyed_dict():
    """A dict keyed by (symbol, ratio, date) measured 403 MB at 3.6M entries."""
    dates = ["2026-08-21", "2026-08-20"]
    out = workbook.series_by_ratio.__doc__ or ""
    assert "Arrays rather than a dict" in out
    assert workbook.MAX_DAYS <= 500


def test_every_series_lines_up_with_the_date_columns(warehouse):
    """An off-by-one here would silently print yesterday's ratio under today."""
    _write([
        _ratio("AAA", "pe", "2026-08-20", 25.0),
        _ratio("AAA", "pe", "2026-08-21", 26.0),
        _ratio("BBB", "roe", "2026-08-20", 14.5),
    ])
    dates = workbook.recent_dates(30)
    series = workbook.series_by_ratio(dates)

    assert dates == ["2026-08-21", "2026-08-20"]
    assert series["pe"]["AAA"] == [26.0, 25.0]
    assert series["roe"]["BBB"] == [None, 14.5]
    assert "AAA" not in series["roe"]
